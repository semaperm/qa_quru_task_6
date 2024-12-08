import csv

from zipfile import ZipFile
from pypdf import PdfReader

from io import BytesIO, StringIO
from openpyxl import load_workbook




def test_pdf_file():
    with ZipFile("new.zip", 'r') as zip_archive:
        pdf_file_archived = BytesIO(zip_archive.read("file_1.pdf"))
        content = PdfReader(pdf_file_archived)
        page = content.pages[0]
        assert "Тестовый PDF-документ" in page.extract_text(0)


def test_xlsx_file():
    with ZipFile("new.zip", 'r') as zip_archive:
        xslx_file = load_workbook(zip_archive.open("file_2.xlsx"))
        sheet = xslx_file['TemplateImportEmpl']
        assert sheet.cell(row=2, column=1).value == "CN001"


def test_csv_file():
    with ZipFile("new.zip", 'r') as zip_archive:
        data = StringIO(zip_archive.read("file_3.csv").decode('utf-8'))
        data_converted = csv.reader(data, delimiter=',')
        list(zip(*data_converted))
        for _ in data_converted:
            for row in data_converted:
                assert "Дорожка на стол Веселые снеговики (50х150см)" in row